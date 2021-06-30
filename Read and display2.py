import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QCheckBox, QLabel, QTableWidget, QTableWidgetItem, QGridLayout, QFileDialog, QDesktopWidget
from pathlib import Path
import string
import openpyxl
class Form(QWidget):
	def __init__(self):
		super().__init__()
		self.initUI()
	def initUI(self):
		self.resize(800, 600)
		self.moveToCenter()
		self.setWindowTitle("Data display")
		grid = QGridLayout()
		self.dataTable = QTableWidget(self)
		self.ROWS = 0
		self.COLUMNS = 0
		self.rows_list = []
		self.columns_list = []
		self.excelContent = None
		chooseFileButton = QPushButton("Choose file", self)
		chooseFileButton.clicked.connect(self.chooseFile)
		self.rowCheckBox = QCheckBox("Row", self)
		self.rowCheckBox.stateChanged.connect(self.contentOnlyStatusChanged)
		self.columnCheckBox = QCheckBox("Column", self)
		self.columnCheckBox.stateChanged.connect(self.contentOnlyStatusChanged)
		self.fileNameLabel = QLabel("File name", self)
		self.contentOnlyLabel = QLabel("Content only")
		grid.addWidget(self.rowCheckBox, 1, 4, 1, 1)
		grid.addWidget(self.columnCheckBox, 2, 4, 1, 1)
		grid.addWidget(self.dataTable, 3, 6, 1, 1)
		grid.addWidget(self.contentOnlyLabel, 0, 4, 1, 1)
		grid.addWidget(chooseFileButton, 4, 4, 1, 1)
		grid.addWidget(self.fileNameLabel, 2, 6, 1, 1)
		self.setLayout(grid)
		self.show()
	def chooseFile(self):
		user = str(Path.home()).removeprefix("C:\\Users\\")
		filename = QFileDialog.getOpenFileName(self, "Choose a file to read data from", f"C:\\Users\\{user}", "*.xlsx")
		if filename[0]:
			FILE_DIR = filename[0]
			self.excelContent = openpyxl.load_workbook(FILE_DIR)
			self.excelContent = self.excelContent.active
			self.ROWS = self.excelContent.max_row
			self.COLUMNS = self.excelContent.max_column
			self.setTableCounts()
			if not self.rowCheckBox.isChecked():
				self.rows_list = []
				for i in range(1, self.ROWS+1):
					row = self.excelContent.cell(i, 1)
					self.rows_list.append(row.value)
			else:
				self.rows_list = []
				self.rows_list = [str(i) for i in range(1, self.ROWS+1)]
			if not self.columnCheckBox.isChecked():
				self.columns_list = []
				for i in range(1, self.COLUMNS+1):
					col = self.excelContent.cell(1, i)
					self.columns_list.append(col.value)
			else:
				self.columns_list = []
				self.columns_list = [str(i) for i in string.ascii_uppercase]
			self.setTableContent()
			self.fileNameLabel.setText(str(FILE_DIR[::-1][:FILE_DIR[::-1].index("/"):][::-1]))
	def contentOnlyStatusChanged(self):
		if self.excelContent != None:
			if not self.rowCheckBox.isChecked():
				self.rows_list = []
				for i in range(1, self.ROWS+1):
					row = self.excelContent.cell(i, 1)
					self.rows_list.append(row.value)
			else:
				self.rows_list = []
				self.rows_list = [str(i) for i in range(1, self.ROWS+1)]
			if not self.columnCheckBox.isChecked():
				self.columns_list = []
				for i in range(1, self.COLUMNS+1):
					col = self.excelContent.cell(1, i)
					self.columns_list.append(col.value)
			else:
				self.columns_list = []
				self.columns_list = [str(i) for i in string.ascii_uppercase]
			self.setTableContent()
	def setTableCounts(self):
		self.dataTable.setRowCount(self.ROWS)
		self.dataTable.setColumnCount(self.COLUMNS)
	def setTableContent(self):
		self.dataTable.setVerticalHeaderLabels(self.rows_list)
		self.dataTable.setHorizontalHeaderLabels(self.columns_list)
		self.fillTableCells()
	def fillTableCells(self):
		for j in range(1, self.COLUMNS + 1):
			for i in range(2, self.ROWS + 1):
				cell = self.excelContent.cell(i, j)
				self.dataTable.setItem(i - 1, j - 1, QTableWidgetItem(str(cell.value)))
	def moveToCenter(self):
		qr = self.frameGeometry()
		cp = QDesktopWidget().availableGeometry().center()
		qr.moveCenter(cp)
		self.move(qr.topLeft())
app = QApplication(sys.argv)
form = Form()
sys.exit(app.exec_())