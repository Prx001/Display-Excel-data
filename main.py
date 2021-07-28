import sys
from PyQt5.QtCore import QSortFilterProxyModel
from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QTableView, QLabel, QLineEdit, QSpinBox, QAction, \
	QFileDialog, QHBoxLayout, QVBoxLayout, QDesktopWidget
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from pathlib import Path
import string
import openpyxl


class MainWidget(QWidget):
	def __init__(self):
		super().__init__()

		self.item_model = QStandardItemModel()
		self.table = QTableView()
		self.filter = QSortFilterProxyModel()
		self.column_label = QLabel("Column:")
		self.spinbox = QSpinBox()
		self.search_label = QLabel("Search:")
		self.search_line_edit = QLineEdit()

		self.initUI()

	def initUI(self):
		hbox = QHBoxLayout()
		vbox = QVBoxLayout()
		hbox.addWidget(self.column_label)
		hbox.addWidget(self.spinbox)
		hbox.addWidget(self.search_label)
		hbox.addWidget(self.search_line_edit)
		vbox.addLayout(hbox)
		vbox.addWidget(self.table)
		self.setLayout(vbox)


class Form(QMainWindow):
	def __init__(self):
		super().__init__()

		self.ROWS = 0
		self.COLUMNS = 0
		self.rows_list = []
		self.columns_list = []
		self.excel_content = None

		self.initUI()

	def initUI(self):
		self.resize(800, 600)
		self.move_to_center()
		self.setWindowTitle("Data display")
		self.central_widget = MainWidget()
		self.central_widget.search_line_edit.textChanged.connect(self.central_widget.filter.setFilterRegExp)
		menu_bar = self.menuBar()
		file_menu = menu_bar.addMenu("File")
		open_action = QAction("Open", self)
		open_action.setShortcut("Ctrl+Shift+O")
		open_action.triggered.connect(self.choose_file)
		file_menu.addAction(open_action)
		self.setCentralWidget(self.central_widget)
		self.show()

	def choose_file(self):
		user = str(Path.home()).removeprefix("C:\\Users\\")
		file_name = QFileDialog.getOpenFileName(self, "Choose a file to read data from", f"C:\\Users\\{user}", "*.xlsx")
		if file_name[0]:
			file_dir = file_name[0]
			self.excel_content = openpyxl.load_workbook(file_dir)
			self.excel_content = self.excel_content.active
			self.ROWS = self.excel_content.max_row
			self.COLUMNS = self.excel_content.max_column
			self.set_table_counts()
			self.set_table_content()
			self.statusBar().showMessage(str(file_dir[::-1][:file_dir[::-1].index("/"):][::-1]))

	def set_table_counts(self):
		self.central_widget.item_model.setRowCount(self.ROWS)
		self.central_widget.item_model.setColumnCount(self.COLUMNS)

	def set_table_content(self):
		self.central_widget.item_model.setVerticalHeaderLabels([str(i) for i in range(1, self.ROWS + 1)])
		self.central_widget.item_model.setHorizontalHeaderLabels([str(i) for i in string.ascii_uppercase])
		self.fill_table_cells()

	def fill_table_cells(self):
		for j in range(1, self.COLUMNS + 1):
			for i in range(1, self.ROWS + 1):
				cell = self.excel_content.cell(i, j)
				self.central_widget.item_model.setItem(i - 1, j - 1, QStandardItem(str(cell.value)))
		self.central_widget.filter.setSourceModel(self.central_widget.item_model)
		self.central_widget.filter.setFilterKeyColumn(self.central_widget.spinbox.value() - 1)
		self.central_widget.table.setModel(self.central_widget.filter)

	def move_to_center(self):
		qr = self.frameGeometry()
		cp = QDesktopWidget().availableGeometry().center()
		qr.moveCenter(cp)
		self.move(qr.topLeft())


app = QApplication(sys.argv)
form = Form()
sys.exit(app.exec_())
