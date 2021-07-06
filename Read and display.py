import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QAction, QFileDialog, \
	QDesktopWidget
from pathlib import Path
import string
import openpyxl


class Form(QMainWindow):
	def __init__(self):
		super().__init__()

		self.data_table = QTableWidget(self)
		self.ROWS = 0
		self.COLUMNS = 0
		self.rows_list = []
		self.columns_list = []
		self.excel_content = None

		self.initUI()

	def initUI(self):
		self.resize(800, 600)
		self.move_to_center()
		self.setWindowTitle("Data display")
		menu_bar = self.menuBar()
		file_menu = menu_bar.addMenu("File")
		open_action = QAction("Open", self)
		open_action.setShortcut("Ctrl+Shift+O")
		open_action.triggered.connect(self.choose_file)
		file_menu.addAction(open_action)
		self.setCentralWidget(self.data_table)
		self.show()

	def choose_file(self):
		user = str(Path.home()).removeprefix("C:\\Users\\")
		file_name = QFileDialog.getOpenFileName(self, "Choose a file to read data from", f"C:\\Users\\{user}", "*.xlsx")
		if file_name[0]:
			file_dir = file_name[0]
			self.excel_content = openpyxl.load_workbook(file_dir)
			self.excel_content = self.excel_content.active
			self.ROWS = self.excel_content.max_row
			self.COLUMNS = self.excel_content.max_column
			self.set_table_counts()
			self.set_table_content()
			self.statusBar().showMessage(str(file_dir[::-1][:file_dir[::-1].index("/"):][::-1]))

	def set_table_counts(self):
		self.data_table.setRowCount(self.ROWS)
		self.data_table.setColumnCount(self.COLUMNS)

	def set_table_content(self):
		self.data_table.setVerticalHeaderLabels([str(i) for i in range(1, self.ROWS + 1)])
		self.data_table.setHorizontalHeaderLabels([str(i) for i in string.ascii_uppercase])
		self.fill_table_cells()

	def fill_table_cells(self):
		for j in range(1, self.COLUMNS + 1):
			for i in range(1, self.ROWS + 1):
				cell = self.excel_content.cell(i, j)
				self.data_table.setItem(i - 1, j - 1, QTableWidgetItem(str(cell.value)))

	def move_to_center(self):
		qr = self.frameGeometry()
		cp = QDesktopWidget().availableGeometry().center()
		qr.moveCenter(cp)
		self.move(qr.topLeft())


app = QApplication(sys.argv)
form = Form()
sys.exit(app.exec_())